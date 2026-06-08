import io
from datetime import datetime, timezone, timedelta, date, time
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config import supabase
from schemas import EmpresaCreate, SedeCreate, UsuarioCreate, PasswordUpdate
from deps import solo_admin
from utils import ahora_iso, fecha_local_a_utc

router = APIRouter(prefix="/api/admin", tags=["admin"])


@router.get("/empresas")
async def listar_empresas(admin: dict = Depends(solo_admin)):
    return supabase.table("empresas").select("id, nombre, nit, direccion, telefono, email, creado_en").order("nombre").execute().data


@router.post("/empresas", status_code=201)
async def crear_empresa(data: EmpresaCreate, admin: dict = Depends(solo_admin)):
    row = data.model_dump()
    row["creado_en"] = ahora_iso()
    return supabase.table("empresas").insert(row).execute().data[0]


@router.put("/empresas/{empresa_id}")
async def editar_empresa(empresa_id: int, data: EmpresaCreate, admin: dict = Depends(solo_admin)):
    result = supabase.table("empresas").update(data.model_dump()).eq("id", empresa_id).execute()
    if not result.data:
        raise HTTPException(404, "Empresa no encontrada")
    return result.data[0]


@router.delete("/empresas/{empresa_id}")
async def eliminar_empresa(empresa_id: int, admin: dict = Depends(solo_admin)):
    supabase.table("empresas").delete().eq("id", empresa_id).execute()
    return {"ok": True}


@router.get("/sedes")
async def listar_sedes(
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    search: Optional[str] = None,
    empresa_id: Optional[int] = None,
    admin: dict = Depends(solo_admin),
):
    q = supabase.table("sedes").select("id, empresa_id, nombre, direccion, ciudad, creado_en, empresas(nombre)", count="exact").order("nombre")
    if search:
        q = q.or_(f"nombre.ilike.%{search}%,direccion.ilike.%{search}%")
    if empresa_id:
        q = q.eq("empresa_id", empresa_id)
    result = q.range(offset, offset + limit - 1).execute()
    return {"items": result.data, "total": result.count}


@router.post("/sedes", status_code=201)
async def crear_sede(data: SedeCreate, admin: dict = Depends(solo_admin)):
    row = data.model_dump()
    row["creado_en"] = ahora_iso()
    return supabase.table("sedes").insert(row).execute().data[0]


@router.put("/sedes/{sede_id}")
async def editar_sede(sede_id: int, data: SedeCreate, admin: dict = Depends(solo_admin)):
    result = supabase.table("sedes").update(data.model_dump()).eq("id", sede_id).execute()
    if not result.data:
        raise HTTPException(404, "Sede no encontrada")
    return result.data[0]


@router.delete("/sedes/{sede_id}")
async def eliminar_sede(sede_id: int, admin: dict = Depends(solo_admin)):
    supabase.table("sedes").delete().eq("id", sede_id).execute()
    return {"ok": True}


@router.get("/usuarios")
async def listar_usuarios(admin: dict = Depends(solo_admin), solo_tecnicos: bool = False):
    if solo_tecnicos:
        return supabase.table("perfiles").select("id, nombre_completo, username, email, rol, activo, creado_en").eq("rol", "tecnico").order("nombre_completo").execute().data
    return supabase.table("perfiles").select("id, nombre_completo, username, email, rol, activo, creado_en").order("nombre_completo").execute().data


@router.post("/usuarios", status_code=201)
async def crear_usuario(data: UsuarioCreate, admin: dict = Depends(solo_admin)):
    try:
        email = f"{data.username}@gmp.com"
        auth_resp = supabase.auth.admin.create_user({
            "email": email,
            "password": data.password,
            "email_confirm": True,
            "user_metadata": {"nombre_completo": data.nombre_completo, "rol": data.rol, "username": data.username},
        })
    except Exception as e:
        raise HTTPException(500, f"Error al crear usuario en Auth: {str(e)}")

    if not auth_resp.user:
        raise HTTPException(500, "Error al crear usuario en Auth: sin respuesta del servidor")

    user_id = auth_resp.user.id
    try:
        supabase.table("perfiles").upsert({
            "id": user_id,
            "email": email,
            "nombre_completo": data.nombre_completo,
            "rol": data.rol,
            "username": data.username,
            "activo": True,
            "creado_en": ahora_iso(),
        }).execute()
    except Exception as e:
        import logging
        logging.warning(f"Perfil upsert failed for user {user_id}: {e}")

    return {"id": user_id, "username": data.username, "rol": data.rol, "nombre_completo": data.nombre_completo}


@router.put("/usuarios/{user_id}")
async def editar_usuario(user_id: str, data: UsuarioCreate, admin: dict = Depends(solo_admin)):
    supabase.table("perfiles").update({
        "nombre_completo": data.nombre_completo,
        "rol": data.rol,
    }).eq("id", user_id).execute()
    return {"ok": True}


@router.put("/usuarios/{user_id}/toggle")
async def toggle_usuario(user_id: str, admin: dict = Depends(solo_admin)):
    perfil = supabase.table("perfiles").select("activo").eq("id", user_id).execute()
    if not perfil.data:
        raise HTTPException(404, "Usuario no encontrado")
    nuevo = not perfil.data[0]["activo"]
    supabase.table("perfiles").update({"activo": nuevo}).eq("id", user_id).execute()
    return {"activo": nuevo}


@router.put("/usuarios/{user_id}/password")
async def cambiar_password(user_id: str, data: PasswordUpdate, admin: dict = Depends(solo_admin)):
    try:
        supabase.auth.admin.update_user_by_id(user_id, {"password": data.password})
    except Exception as e:
        raise HTTPException(500, f"Error al cambiar contrasena: {str(e)}")
    return {"ok": True}


# ─── Dashboard ────────────────────────────────────────────────────────────────


def _build_date_filters(fecha_inicio: Optional[str], fecha_fin: Optional[str]):
    """Returns (gte, lte) tuples for Supabase date filtering."""
    gte = None
    lte = None
    if fecha_inicio:
        gte = fecha_local_a_utc(fecha_inicio, inicio=True)
    if fecha_fin:
        lte = fecha_local_a_utc(fecha_fin, inicio=False)
    return gte, lte


@router.get("/dashboard/resumen")
async def dashboard_resumen(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    empresa_id: Optional[int] = None,
    admin: dict = Depends(solo_admin),
):
    gte, lte = _build_date_filters(fecha_inicio, fecha_fin)

    params = {}
    if gte:
        params["p_fecha_inicio"] = gte
    if lte:
        params["p_fecha_fin"] = lte
    if empresa_id:
        params["p_empresa_id"] = empresa_id

    result = supabase.rpc("dashboard_resumen", params).execute()

    if result.data:
        return result.data
    return {
        "total_visitas": 0,
        "visitas_por_tecnico": [],
        "puntos_mas_visitados": [],
    }


@router.get("/dashboard/puntuaciones")
async def dashboard_puntuaciones(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    admin: dict = Depends(solo_admin),
):
    gte, lte = _build_date_filters(fecha_inicio, fecha_fin)

    params = {}
    if gte:
        params["p_fecha_inicio"] = gte
    if lte:
        params["p_fecha_fin"] = lte

    result = supabase.rpc("dashboard_puntuaciones", params).execute()

    if result.data:
        return result.data
    return {"puntuaciones": []}


@router.get("/reportes/exportar-excel")
async def exportar_reportes_excel(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    admin: dict = Depends(solo_admin),
):
    q = (
        supabase.from_("vista_reportes_busqueda")
        .select("id, fecha_hora, empresa_nombre, sede_nombre, tecnico_nombre, motivo_visita, motivo_visita_otro, nombre_asesor, telefono_asesor, hallazgos, uso_materiales, materiales_detalle, cambio_antena, serial_antena")
        .order("fecha_hora", desc=True)
    )
    gte, lte = _build_date_filters(fecha_inicio, fecha_fin)
    if gte:
        q = q.gte("fecha_hora", gte)
    if lte:
        q = q.lte("fecha_hora", lte)
    rows = q.execute().data

    wb = Workbook()
    ws = wb.active
    ws.title = "Reportes"

    headers = [
        "ID", "Fecha", "Hora", "Empresa", "Sede", "Tecnico", "Motivo Visita",
        "Motivo Visita Otro", "Asesor", "Telefono Asesor", "Hallazgos",
        "Uso Materiales", "Materiales Detalle", "Cambio Antena", "Serial Antena",
    ]

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="131313", end_color="131313", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    cell_font = Font(name="Calibri", size=11)
    cell_align = Alignment(vertical="center", wrap_text=True)
    date_fmt = "DD/MM/YYYY"
    time_fmt = "HH:MM"
    tz_local = timezone(timedelta(hours=-5))

    for row_idx, r in enumerate(rows, 2):
        motivo = r.get("motivo_visita", "")
        motivo_otro = r.get("motivo_visita_otro", "") or ""
        if motivo == "otro" and motivo_otro:
            motivo_display = f"Otro: {motivo_otro}"
        else:
            motivo_display = motivo

        raw_dt = r.get("fecha_hora", "")
        fecha_val = None
        hora_val = None
        if raw_dt:
            try:
                dt = datetime.fromisoformat(str(raw_dt).replace("Z", "+00:00")).astimezone(tz_local)
                fecha_val = dt.date()
                hora_val = dt.time()
            except Exception:
                fecha_val = raw_dt
                hora_val = ""

        values = [
            r.get("id", ""),
            fecha_val,
            hora_val,
            r.get("empresa_nombre", ""),
            r.get("sede_nombre", ""),
            r.get("tecnico_nombre", ""),
            motivo_display,
            motivo_otro if motivo == "otro" else "",
            r.get("nombre_asesor", ""),
            r.get("telefono_asesor", ""),
            r.get("hallazgos", ""),
            "Si" if r.get("uso_materiales") else "No",
            r.get("materiales_detalle", ""),
            "Si" if r.get("cambio_antena") else "No",
            r.get("serial_antena", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            if col_idx == 2 and isinstance(val, date):
                cell.number_format = date_fmt
            elif col_idx == 3 and isinstance(val, time):
                cell.number_format = time_fmt

    column_widths = [6, 12, 8, 20, 20, 22, 16, 18, 20, 16, 30, 14, 30, 14, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reportes_{ahora_iso()[:10]}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
